from PyPDF2 import PdfReader
import csv
import pandas as pd

# Function to extract student names from PDF
def extract_students_from_pdf(pdf_path):
    """Extract student data from PDF, filtering out headers and page markers"""
    student_list = []
    reader = PdfReader(pdf_path)

    # Skip the first page (index 0), start from page 1
    for page in reader.pages[1:]:
        text = page.extract_text()
        lines = text.split("\n")
        
        for line in lines:
            line = line.strip()
            
            # Skip empty lines
            if not line:
                continue
            
            # Skip column headers
            if 'NO.' in line and 'STUDENT' in line and 'SURNAME' in line:
                continue
            
            # Skip university name and degree info
            if any(x in line for x in ['THE UNIVERSITY', 'BACHELOR', 'GENDER']):
                continue
            
            # Skip single digit page numbers
            if line.isdigit() and 1 <= int(line) <= 50:
                continue
            
            # Skip status markers
            if line.lower() in ['regular', 'male', 'female']:
                continue
            
            student_list.append(line)

    return student_list

# Function to search for a student by name
def find_student(student_list, name):
    for line in student_list:
        if name.upper() in line.upper():  # Case insensitive search
            return line.strip()
    return None

# Function to parse student data into separate columns with intelligent pattern matching
def parse_student_data_columns(student_list):
    """Parse student records into separate columns with robust pattern matching for mixed data formats"""
    parsed_data = []
    
    for line in student_list:
        line = line.strip()
        if not line:
            continue
        
        # Skip header and separator lines
        if any(x in line for x in ['NO.', 'STUDENT', 'GENDER', 'THE UNIVERSITY', 'BACHELOR', 'REGULAR', 'MALE', 'FEMALE']):
            if not any(c.isdigit() for c in line.split()[0]):  # Skip if first token isn't numeric
                continue
        
        # Split the line by spaces
        parts = line.split()
        
        # Need at least 3 parts: could be NO+STUD_NO combined, or separate
        if len(parts) < 2:
            continue
        
        try:
            no = ''
            student_no = ''
            remaining_parts = []
            
            # Check if first part is NO. and STUDENT NO. combined (e.g., "322016138632")
            first_part = parts[0]
            
            if len(first_part) > 2 and first_part.isdigit() and len(first_part) >= 10:
                # Likely combined NO. and STUDENT NO.
                # Try different split positions (NO is typically 1-3 digits, STUDENT NO is 10 digits)
                for split_pos in range(1, 4):  # Try positions 1-3 for NO length
                    potential_no = first_part[:split_pos]
                    potential_stud_no = first_part[split_pos:]
                    if len(potential_stud_no) == 10 and potential_stud_no.isdigit():
                        no = potential_no
                        student_no = potential_stud_no
                        remaining_parts = parts[1:]
                        break
                
                # If still no match, treat as separate fields
                if not no and not student_no:
                    if first_part.isdigit() and len(first_part) == 10:
                        no = ''  # Will skip this record
                    continue
            elif first_part.isdigit() and len(parts) > 1:
                # NO. and STUDENT NO. are separate
                no = first_part
                if parts[1].isdigit() and len(parts[1]) == 10:
                    student_no = parts[1]
                    remaining_parts = parts[2:]
                else:
                    continue
            else:
                continue
            
            # Now we have NO., STUDENT NO., and remaining parts are SURNAME and NAMES
            if not no or not student_no or len(remaining_parts) < 1:
                continue
            
            surname = remaining_parts[0]
            
            # Identify NRC/PASSPORT - typically contains "/" or is a long number at the end
            nrc_passport = ''
            name_parts = []
            
            # Check from the end to find NRC/PASSPORT
            for i in range(len(remaining_parts) - 1, 0, -1):
                part = remaining_parts[i]
                # NRC/PASSPORT pattern: contains "/" or is 9+ digits
                if '/' in part or (len(part) >= 9 and part.isdigit()):
                    nrc_passport = part
                    name_parts = remaining_parts[1:i]  # Everything between surname and NRC/PASSPORT
                    break
            
            # If no NRC/PASSPORT found, all parts after surname are names
            if not nrc_passport and len(remaining_parts) > 1:
                name_parts = remaining_parts[1:]
            
            # Parse names: FIRST NAME is first, rest are OTHER NAMES
            first_name = ''
            other_names = ''
            
            if len(name_parts) >= 1:
                first_name = name_parts[0]
                if len(name_parts) > 1:
                    other_names = ' '.join(name_parts[1:])
            
            # Add record if we have the required fields
            if no and student_no and surname:
                parsed_data.append({
                    'NO.': no,
                    'STUDENT NO.': student_no,
                    'SURNAME': surname,
                    'OTHER NAMES': other_names,
                    'FIRST NAME': first_name,
                    'NRC/PASSPORT': nrc_passport
                })
                
        except (IndexError, ValueError) as e:
            continue
    
    return parsed_data

# Function to save student data to Excel with separate columns and professional formatting
def save_students_to_excel_columns(parsed_data, excel_path):
    """Save the extracted student list to an Excel file with separate columns and professional formatting"""
    
    if not parsed_data:
        print("No valid student data to save.")
        return
    
    # Create a DataFrame from the parsed data
    df = pd.DataFrame(parsed_data)

    # Save to Excel file
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Student List', index=False, startrow=2)

        # Get workbook and worksheet
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        worksheet = writer.sheets['Student List']

        # Add title at the top
        worksheet['A1'] = 'STUDENT ENROLLMENT LIST'
        worksheet.merge_cells('A1:F1')
        
        # Style the title row
        title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        title_cell = worksheet['A1']
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        
        # Set title row height
        worksheet.row_dimensions[1].height = 30

        # Define column widths based on content
        column_widths = {
            'A': 8,   # NO.
            'B': 14,  # STUDENT NO.
            'C': 16,  # SURNAME
            'D': 20,  # OTHER NAMES
            'E': 14,  # FIRST NAME
            'F': 18   # NRC/PASSPORT
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define borders
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Style the header row (now in row 3 due to title)
        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f'{col_letter}3']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set header row height
        worksheet.row_dimensions[3].height = 25

        # Style data rows (starting from row 4)
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        # Apply styling to all data rows
        for row_num in range(4, len(parsed_data) + 4):
            worksheet.row_dimensions[row_num].height = 20
            for col_num in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f'{col_letter}{row_num}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # Freeze the title and header rows
        worksheet.freeze_panes = 'A4'

        # Add autofilter to header
        worksheet.auto_filter.ref = f'A3:F{len(parsed_data)+3}'

    print(f"Student data saved to {excel_path} with separate columns and professional formatting")
    print(f"Total records: {len(parsed_data)}")

# Main execution
pdf_path = "/home/paul/Documents/Google CyberSecurity Scholarship/Cyber security Portfolio Project /Google-Cybersecurity-Professional-Portfolio/07-Automate-Cybersecurity-Tasks-with-Python/hss_(b.a_nqs)_final_pub._list_-_male.pdf"
excel_path = "/home/paul/Documents/Google CyberSecurity Scholarship/Cyber security Portfolio Project /Google-Cybersecurity-Professional-Portfolio/07-Automate-Cybersecurity-Tasks-with-Python/students_list.xlsx"

# Step 1: Extract student data from PDF
print("Step 1: Extracting student data from PDF...")
student_list = extract_students_from_pdf(pdf_path)
print(f"✓ Extracted {len(student_list)} total student records from PDF")

# Step 2: Parse student data into separate columns
print("\nStep 2: Parsing student data into columns...")
parsed_data = parse_student_data_columns(student_list)
print(f"✓ Parsed {len(parsed_data)} student records into structured format")

# Step 3: Save all students to Excel file with separate columns
print("\nStep 3: Saving parsed data to Excel file...")
save_students_to_excel_columns(parsed_data, excel_path)

# Step 4: Search for a specific student
print("\nStep 4: Searching for student...")
student_name = "PAUL"  # Adjusted to match the PDF format
found = find_student(student_list, student_name)

if found:
    print(f"✓ Found student: {found}")
else:
    print(f"✗ Student '{student_name}' not found in the list.")

# Summary
print("\n" + "="*60)
print("PROCESS COMPLETED SUCCESSFULLY")
print("="*60)
print(f"Total students extracted from PDF: {len(student_list)}")
print(f"Total records parsed and saved: {len(parsed_data)}")
print(f"Excel file created: {excel_path.split('/')[-1]}")
print("="*60)